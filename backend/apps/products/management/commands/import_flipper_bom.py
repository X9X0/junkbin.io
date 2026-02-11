"""
Management command to import Flipper Zero BOM data from Excel files.

Creates Component records and ProductComponent links for all parts in the
Main PCB and iButton/NFC PCB BOMs. Also generates a comprehensive Excel
spreadsheet for download.

Usage:
    python manage.py import_flipper_bom
    python manage.py import_flipper_bom --flush        # Delete existing links and re-import
    python manage.py import_flipper_bom --dry-run       # Show what would be imported
    python manage.py import_flipper_bom --spreadsheet   # Only generate spreadsheet
    python manage.py import_flipper_bom --designators   # Apply reference designators from schematics
"""
import os
import re
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.db import transaction

User = get_user_model()

# BOM column indices (row 6 has headers)
COL_MOUSER = 0
COL_MFR_PN = 1
COL_MFR = 2
COL_CUST_PN = 3
COL_DESC = 4
COL_ROHS = 5
COL_LIFECYCLE = 6
COL_QTY = 7
COL_PRICE = 8
COL_TOTAL = 9

# Header row index (0-indexed)
HEADER_ROW = 6
DATA_START = 7

BOM_FILES = [
    ("flipper-bom-main-pcb.xls", "Main PCB"),
    ("flipper-bom-ibutton-pcb.xls", "iButton/NFC PCB"),
]

# Reference designator mapping extracted from Flipper Zero schematics.
# Source: docs.flipper.net/zero/development/hardware/schematic
# Main PCB: Rev 11.F7B9C6 (August 2021), iButton/NFC PCB: Rev March 2022
# Note: BOM revision may differ slightly. Only confident matches included.
# Value: string (same for all boards) OR dict {board_name: designators}
REFERENCE_DESIGNATORS = {
    # ===== Main PCB — ICs & Active =====
    'STM32WB55RGV6TR': 'DD1',                  # MCU
    'STM32WB55RGV6': 'DD1',                    # MCU (seed data PN)
    'CC1101RGPR': 'DD2',                        # Sub-1 GHz transceiver
    'BQ27220YZFR': 'DD5',                       # Battery fuel gauge
    'BQ25896RTWR': 'DD6',                       # Li-Po charger
    'BQ25896': 'DD6',                            # (seed data PN)
    'LM3281YFQR': 'DD7, DA6',                   # DC-DC converters (x2)
    'LP5562TMX/NOPB': 'DA4',                    # LED driver
    'ADP151ACPZ-3.3-R7': 'DA3',                 # LDO regulator
    'NC7SZ125M5X': 'DD8',                       # Buffer gate

    # ===== Main PCB — RF =====
    'SKY13373-460LF': 'SW1, SW2',               # RF switches (x2)
    'LFL182G45TC1A108': 'FL1',                   # BLE bandpass filter
    'B0310J50100AHF': 'FL2',                     # Balun

    # ===== Main PCB — Connectors =====
    '105450-0101': 'X1',                         # USB Type-C
    'DM3AT-SF-PEJM5': 'X7',                     # microSD card slot
    'FH34SRJ-30S-0.5SH(50)': 'X3',              # LCD FPC (30-pin)
    'FH34SRJ-18S-0.5SH(50)': 'X6',              # iButton PCB FPC (18-pin)

    # ===== Main PCB — Crystals =====
    'FC-135 32.7680KA-A0': 'ZQ1',                # 32.768 kHz (RTC)
    'ABM3B-32.000MHZ-B2-T': 'ZQ2',              # 32 MHz (MCU HSE)
    'XRCGB26M000F1SABR0': 'ZQ3',                # 26 MHz (CC1101)

    # ===== Main PCB — Discrete Semiconductors =====
    'DMP1045U-7': 'VT1',                         # P-ch MOSFET (power)
    'MMDT2907A-7-F': 'VT2',                      # Dual PNP transistor
    'DMN313DLT-7': 'VT5',                        # N-ch MOSFET (power)
    '1N4148X-TP': 'VD16, VD28, VD31, VD32',      # Signal diodes
    'CLMUD-FKC-CHJNPM569ABB845343': 'HL1',       # RGB LED

    # ===== Main PCB — Misc =====
    '1001312': 'ANT1',                            # BLE chip antenna
    'SKRMAAE010': 'SW3, SW4, SW5, SW6, SW7, SW8',  # Tactile buttons (x6)
    'ERB-RD2R00X': 'FU4',                        # Fuse
    'ST7567': 'HG1',                              # LCD display (seed data)

    # ===== Main PCB — Resistors =====
    'AC0402JR-075K1L': 'R1, R2',                  # 5.1kΩ (USB CC)
    'RC0402JR-0722RP': 'R8, R9',                  # 22Ω
    'MSMA2512R0100FEN': 'R37',                    # 0.01Ω current sense
    'CRCW0402470RFKEDC': 'R36',                   # 470Ω
    'CR0402-FX-5100GLF': 'R63',                   # 510Ω
    'CR0402-FX-3012GLF': 'R39',                   # 30.1kΩ
    'RC0402DR-07100RL': 'R45',                    # 100Ω
    'RC0402JR-7D1KL': 'R69',                      # 1kΩ
    'RT0402FRE0756KL': 'R13',                     # 56kΩ
    'RC0402JR-7D10KL': 'R34, R35, R44, R46, R47, R49, R50, R51, R53, R60, R62',  # 10kΩ (x11)
    'CR0402-JW-331GLF': 'R54, R55, R56, R57, R58, R59',  # 330Ω (buttons)
    'CR0402-JW-510GLF': 'R7, R20, R29',           # 51Ω
    'ERJ-2RKF5231X': 'R38',                       # 5.23kΩ

    # ===== Main PCB — Capacitors =====
    'MSASU105SB7473KFNB25': 'C64',                # 47nF
    '04025A8R2CAT4A': 'C45',                      # 8.2pF
    'GRT0335C1ER30WA02D': 'C6',                   # 0.3pF
    '0402N0R6C500CT': 'C33',                      # 0.6pF
    '0402N0R8C500CT': 'C5',                       # 0.8pF
    'UMK105CH220KV-F': 'C91',                     # 22pF
    'C0402C129C5GACTU': 'C31, C32',               # 1.2pF (x2)
    '0402B101K100CT': {                            # 100pF
        'Main PCB': 'C29, C38',
        'iButton/NFC PCB': 'C1, C8, C10',
    },
    'CC0603ZRY5V6BB105': 'C69',                   # 1µF (0603)
    'C1005JB1C474K050BC': 'C65',                  # 0.47µF

    # ===== Main PCB — Capacitors (bulk, matched by qty + value) =====
    'CL05A475MP5NRNC': 'C15, C19, C20, C73, C77, C81',  # 4.7µF (x6)
    'GRM155R60J106ME05D': 'C67, C68, C71, C72, C83',     # 10µF (x5)

    # ===== Main PCB — Resistors (additional) =====
    'ASC0402-100KFT10': 'R3, R4, R64',                    # 100kΩ
    'EXB-28V510JX': 'R10, R11, R32, R33',                 # 51Ω arrays (x4)

    # ===== Main PCB — Inductors =====
    'LQH44PN1R0MPRL': 'L12',                      # 1µH
    'MLF2012E100JT000': 'L3',                     # 10µH
    'LQG15HS56NJ02D': 'L16',                      # 56nH
    'MCL1005-220-R': 'L8',                         # 22nH
    'LQG15HN3N3C02D': 'L4',                       # 3.3nH
    'LQG15HN6N8H02D': 'L5',                       # 6.8nH
    'LQG15HN2N2C02D': 'L9',                       # 2.2nH
    'MCLA1005V2-3R0-R': 'L1',                     # 3nH (BLE matching)
    'LQG15HN2N7S02D': 'L2, L11',                  # 2.7nH (x2)
    'MCL2012V2-471-R': 'L13, L14',                # 470nH (power)
    'LCCI0402J27NGTAR': 'L19',                    # 27nH

    # ===== iButton/NFC PCB =====
    'ST25R3916-AQWT': 'DD1',                      # NFC transceiver
    'DMN2058UW-7': 'VT2',                         # N-ch MOSFET (vibro)
    'AST0760MCTRQ': 'LS1',                        # Piezo buzzer
    'TSOP75338TR': 'BM1',                          # IR receiver
    '150060BS75000': 'HL2',                        # Blue LED
    '150060GS75000': 'HL3',                        # Green LED
    'LTST-C190KFKT': 'HL1',                       # Orange LED
}


def classify_component(desc):
    """Determine component_type from German Mouser description."""
    dl = desc.lower()
    if 'kondensator' in dl or 'mlcc' in dl:
        return 'capacitor'
    if any(x in dl for x in ['dickfilmwiderst', 'dünnfilmwiderst', 'strommesswiderst']):
        return 'resistor'
    if 'widerstandnetzwerk' in dl:
        return 'resistor'
    if 'induktiv' in dl:
        return 'inductor'
    if 'led' in dl and 'treiber' not in dl:
        return 'led'
    if 'esd-' in dl or 'tvs' in dl:
        return 'diode'
    if 'dioden' in dl:
        return 'diode'
    if 'mosfet' in dl or 'fet ' in dl:
        return 'mosfet'
    if 'transistor' in dl or 'bjt' in dl:
        return 'transistor'
    if any(x in dl for x in ['usb-stecker', 'speicherkarten', 'ffc', 'fpc', 'sockel', 'header']):
        return 'connector'
    if 'quarz' in dl or 'crystal' in dl:
        return 'crystal'
    if 'sicherung' in dl or 'fuse' in dl:
        return 'fuse'
    if 'antenn' in dl:
        return 'antenna'
    if 'sensor' in dl or 'ir sensor' in dl:
        return 'sensor'
    if any(x in dl for x in ['lineare spannungsregler', 'ldo']):
        return 'regulator'
    if any(x in dl for x in ['schaltspannungsregler', 'dc-dc']):
        return 'regulator'
    if 'transceiver' in dl or 'hf-trans' in dl:
        return 'rf_module'
    if 'schalt-ic' in dl:
        return 'ic'
    if any(x in dl for x in ['batterie', 'battery', 'fuel', 'charger']):
        return 'ic'
    if any(x in dl for x in ['treiber', 'driver', 'puffer', 'buffer']):
        return 'ic'
    return 'other'


def extract_package(desc):
    """Extract package type from description."""
    for pkg in ['2512', '0805', '0603', '0402', '0201']:
        if pkg in desc:
            return pkg
    m = re.search(r'(\d+)-DSBGA', desc)
    if m:
        return f'DSBGA-{m.group(1)}'
    m = re.search(r'(\d+)-WQFN', desc)
    if m:
        return f'WQFN-{m.group(1)}'
    if 'QFN' in desc:
        m = re.search(r'QFN-(\d+)', desc)
        if m:
            return f'QFN-{m.group(1)}'
    if 'SOT323' in desc:
        return 'SOT-323'
    if 'SOT-23' in desc:
        return 'SOT-23'
    return ''


def extract_value(desc, ctype):
    """Extract primary electrical value from description."""
    dl = desc.lower()
    if ctype == 'capacitor':
        m = re.search(r'(\d+\.?\d*)\s*(uf|µf|pf|nf)', dl)
        if m:
            val, unit = m.group(1), m.group(2).upper()
            unit = unit.replace('UF', 'µF').replace('PF', 'pF').replace('NF', 'nF')
            return f'{val} {unit}'
        # Bare "F" (e.g., "0.47F") — treat as µF
        m = re.search(r'(\d+\.?\d*)\s*f\b', dl)
        if m:
            return f'{m.group(1)} µF'
    elif ctype == 'resistor':
        # "5.1k Ohm", "56K ohm" (separate k and ohm)
        m = re.search(r'(\d+\.?\d*)\s*k\s*ohm', dl)
        if m:
            return f'{m.group(1)} kΩ'
        m = re.search(r'(\d+\.?\d*)\s*(kohm|kohms|ohm|ohms|mohm)', dl)
        if m:
            val, unit = m.group(1), m.group(2)
            if 'kohm' in unit:
                return f'{val} kΩ'
            if 'mohm' in unit:
                return f'{val} mΩ'
            return f'{val} Ω'
        # EIA notation: "4K7" = 4.7kΩ
        m = re.search(r'\b(\d+)K(\d)\b', desc)
        if m:
            val = int(m.group(1)) + int(m.group(2)) / 10
            return f'{val:g} kΩ'
    elif ctype == 'inductor':
        m = re.search(r'(\d+\.?\d*)\s*(uh|nh|µh)', dl)
        if m:
            val, unit = m.group(1), m.group(2).upper()
            unit = unit.replace('UH', 'µH').replace('NH', 'nH')
            return f'{val} {unit}'
    elif ctype == 'crystal':
        m = re.search(r'(\d+\.?\d*)\s*(mhz|khz|ghz)', dl)
        if m:
            val, unit = m.group(1), m.group(2)
            unit = unit.upper().replace('MHZ', 'MHz').replace('KHZ', 'kHz').replace('GHZ', 'GHz')
            return f'{val} {unit}'
    elif ctype in ('diode', 'mosfet', 'transistor'):
        m = re.search(r'(\d+)\s*v', dl)
        if m:
            return f'{m.group(1)} V'
    elif ctype == 'fuse':
        m = re.search(r'(\d+\.?\d*)\s*a\b', dl)
        if m:
            return f'{m.group(1)} A'
    return ''


def extract_specifications(desc, ctype):
    """Extract structured specifications from description for the Component.specifications JSON field.

    Returns a dict with keys matching what Component.primary_value expects,
    e.g. {"capacitance_pf": 100, "voltage_v": 10} for a capacitor.
    """
    specs = {}
    dl = desc.lower()

    if ctype == 'capacitor':
        # Capacitance: "100pF", "2.2 uF", "0.1uF", "4.7 nF", "0.47F" (bare F = µF)
        m = re.search(r'(\d+\.?\d*)\s*(uf|µf|pf|nf)', dl)
        if m:
            val = float(m.group(1))
            unit = m.group(2).lower().replace('µ', 'u')
            if unit == 'uf':
                specs['capacitance_uf'] = val
            elif unit == 'nf':
                specs['capacitance_nf'] = val
            elif unit == 'pf':
                specs['capacitance_pf'] = val
        else:
            # Bare "F" without prefix (e.g., "0.47F") — treat as µF for MLCCs
            m = re.search(r'(\d+\.?\d*)\s*f\b', dl)
            if m:
                specs['capacitance_uf'] = float(m.group(1))
        # Voltage: "10V", "50V", "6.3 VDC", "25 VDC"
        m = re.search(r'(\d+\.?\d*)\s*v(?:dc)?\b', dl)
        if m:
            specs['voltage_v'] = float(m.group(1))
        # Tolerance: "10%", "5%", "20%"
        m = re.search(r'(\d+)%', desc)
        if m:
            specs['tolerance_pct'] = int(m.group(1))

    elif ctype == 'resistor':
        # Resistance: "150 OHM", "30.1KOHM", "5.1k Ohm", "56K ohm", "10 kOhms"
        m = re.search(r'(\d+\.?\d*)\s*k\s*ohm', dl)
        if m:
            specs['resistance_ohm'] = float(m.group(1)) * 1000
        else:
            m = re.search(r'(\d+\.?\d*)\s*(mohm|kohm|kohms|ohm|ohms)', dl)
            if m:
                val = float(m.group(1))
                unit = m.group(2)
                if 'kohm' in unit:
                    specs['resistance_ohm'] = val * 1000
                elif 'mohm' in unit:
                    specs['resistance_ohm'] = val / 1000
                else:
                    specs['resistance_ohm'] = val
            else:
                # EIA notation: "4K7" = 4.7kΩ, "10K" = 10kΩ, "2M2" = 2.2MΩ
                m = re.search(r'\b(\d+)k(\d)\b', dl)
                if m:
                    specs['resistance_ohm'] = (int(m.group(1)) + int(m.group(2)) / 10) * 1000
                else:
                    m = re.search(r'\b(\d+)k\b', dl)
                    if m:
                        specs['resistance_ohm'] = float(m.group(1)) * 1000
        # Power: "62.5mW", "1/16WATT", "1/16W"
        m = re.search(r'(\d+\.?\d*)\s*mw', dl)
        if m:
            specs['power_w'] = float(m.group(1)) / 1000
        else:
            m = re.search(r'1/16\s*w', dl)
            if m:
                specs['power_w'] = 0.0625
        # Tolerance
        m = re.search(r'(\d+)%', desc)
        if m:
            specs['tolerance_pct'] = int(m.group(1))

    elif ctype == 'inductor':
        # Inductance: "22 NH", "470 NH", "3.0 NH", "27nH", "2.7 nH"
        m = re.search(r'(\d+\.?\d*)\s*(uh|nh|µh)', dl)
        if m:
            val = float(m.group(1))
            unit = m.group(2).lower().replace('µ', 'u')
            if unit == 'uh':
                specs['inductance_uh'] = val
            elif unit == 'nh':
                specs['inductance_nh'] = val

    elif ctype == 'crystal':
        # Frequency: "32.0000MHZ", "32.768KHz", "26.0MHz"
        m = re.search(r'(\d+\.?\d*)\s*(mhz|khz|ghz)', dl)
        if m:
            val = float(m.group(1))
            unit = m.group(2)
            if 'ghz' in unit:
                specs['frequency_mhz'] = val * 1000
            elif 'khz' in unit:
                specs['frequency_mhz'] = val / 1000
            else:
                specs['frequency_mhz'] = val

    elif ctype == 'diode':
        # Voltage: "100Vrm", "24V"
        m = re.search(r'(\d+)\s*v', dl)
        if m:
            specs['vf_v'] = float(m.group(1))
        # Current: "300mA"
        m = re.search(r'(\d+)\s*ma', dl)
        if m:
            specs['current_a'] = float(m.group(1)) / 1000

    elif ctype == 'mosfet':
        # BVDSS voltage: "BVDSS 25V", "BVDSS: 8V-24V"
        m = re.search(r'bvdss[:\s]*(?:\d+v[- ]*)?(\d+)\s*v', dl)
        if m:
            specs['vds_v'] = float(m.group(1))

    elif ctype == 'transistor':
        # Voltage: "-60V" or "60V"
        m = re.search(r'-?(\d+)\s*v', dl)
        if m:
            specs['vce_v'] = float(m.group(1))
        # Power: "200mW"
        m = re.search(r'(\d+)\s*mw', dl)
        if m:
            specs['power_w'] = float(m.group(1)) / 1000

    elif ctype == 'fuse':
        # Current: "2A"
        m = re.search(r'(\d+\.?\d*)\s*a\b', dl)
        if m:
            specs['current_a'] = float(m.group(1))

    elif ctype == 'led':
        # Wavelength: "605nm"
        m = re.search(r'(\d+)\s*nm', dl)
        if m:
            specs['wavelength_nm'] = int(m.group(1))
        # Color detection
        for color in ['Red', 'Green', 'Blue', 'Orange', 'Yellow', 'White', 'RGB']:
            if color.lower() in dl:
                specs['color'] = color
                break

    elif ctype == 'connector':
        # Pin count: "18P", "30P", "02 SIL", "24" (from USB Type C)
        m = re.search(r'(\d+)\s*p(?:os|in)?\b', dl)
        if m:
            specs['pins'] = int(m.group(1))
        elif 'sil' in dl:
            m = re.search(r'(\d+)\s*sil', dl)
            if m:
                specs['pins'] = int(m.group(1))
        # Standard detection
        if 'usb type c' in dl or 'usb-c' in dl:
            specs['standard'] = 'USB Type-C'
        elif 'microsd' in dl or 'micro sd' in dl:
            specs['standard'] = 'microSD'

    elif ctype == 'regulator':
        # Output current: "200 mA", "1.2-A", "3A", "1.2A"
        m = re.search(r'(\d+\.?\d*)\s*-?\s*a\b', dl)
        if m:
            specs['iout_a'] = float(m.group(1))
        m = re.search(r'(\d+)\s*ma\b', dl)
        if m:
            specs['iout_a'] = float(m.group(1)) / 1000
        # Output voltage: "3.3-V", "3.3V"
        m = re.search(r'(\d+\.\d+)\s*-?\s*v\b', dl)
        if m:
            specs['output_v'] = float(m.group(1))

    elif ctype == 'rf_module':
        # Frequency: ".1-6.0GHz", "6.0GHz"
        m = re.search(r'(\d+\.?\d*)\s*ghz', dl)
        if m:
            specs['frequency_ghz'] = float(m.group(1))

    elif ctype == 'ic':
        # Various ICs — try frequency, channels
        m = re.search(r'(\d+\.?\d*)\s*ghz', dl)
        if m:
            specs['frequency_ghz'] = float(m.group(1))
        m = re.search(r'(\d+)\s*-?\s*ch\b', dl)
        if m:
            specs['channels'] = int(m.group(1))
        # Current for battery management: "3A"
        m = re.search(r'(\d+\.?\d*)\s*a\b', dl)
        if m and 'charg' in dl:
            specs['charge_current_a'] = float(m.group(1))

    elif ctype == 'sensor':
        # IR sensors: "38kHz"
        m = re.search(r'(\d+)\s*khz', dl)
        if m:
            specs['frequency_khz'] = int(m.group(1))

    elif ctype == 'other':
        # Signal conditioning: frequency
        m = re.search(r'(\d+\.?\d*)\s*ghz', dl)
        if m:
            specs['frequency_ghz'] = float(m.group(1))
        elif re.search(r'(\d+\.?\d*)\s*mhz', dl):
            m = re.search(r'(\d+\.?\d*)\s*mhz', dl)
            specs['frequency_mhz'] = float(m.group(1))

    return specs


def translate_function(desc, ctype):
    """Create a clean English typical_function from the German description."""
    dl = desc.lower()
    translations = {
        'kondensator aus mehreren keramikschichten': 'Multilayer ceramic capacitor (MLCC)',
        'dickfilmwiderstände': 'Thick film resistor',
        'dünnfilmwiderstände': 'Thin film resistor',
        'widerstandnetzwerke': 'Resistor network',
        'hf-induktivitäten': 'RF inductor',
        'lineare spannungsregler': 'LDO voltage regulator',
        'schaltspannungsregler': 'Switching voltage regulator',
        'hf-transceiver': 'RF transceiver',
        'batterie-management': 'Battery management IC',
        'led-anzeigetreiber': 'LED display driver',
        'puffer u. leitungstreiber': 'Buffer / line driver',
        'signalaufbereitung': 'Signal conditioning filter',
        'esd-entstörer': 'ESD protection diode',
        'dioden': 'General purpose diode',
        'sensorschalter': 'Tactile switch',
        'infrarot-receiver': 'IR receiver module',
        'piezo-summer': 'Piezo buzzer',
        'speicherkartenverbinder': 'microSD card connector',
        'usb-stecker': 'USB Type-C receptacle',
        'ffc & fpc-steckverbinder': 'FFC/FPC connector',
        'sockel & kabelgehäuse': 'Pin header (SMT)',
        'quarz': 'Crystal oscillator',
        'sicherungen': 'SMD fuse',
        'standard-leds': 'LED indicator',
        'hf-schalt-ics': 'RF switch IC',
        'strommesswiderstände': 'Current sense resistor',
        'antennen': 'Chip antenna (WiFi/BT)',
        'bipolartransistoren': 'BJT transistor',
    }
    for german, english in translations.items():
        if german in dl:
            return english
    return desc[:80] if desc else ''


# German→English translations for Mouser category prefixes and terms
GERMAN_TRANSLATIONS = [
    # Longest/most specific first to avoid partial matches
    ('Kondensator aus mehreren Keramikschichten MLCC - SMD/SMT', 'Multilayer Ceramic Capacitor MLCC - SMD/SMT'),
    ('Kondensator aus mehreren Keramikschichten', 'Multilayer Ceramic Capacitor'),
    ('Dickfilmwiderstände - SMD', 'Thick Film Resistors - SMD'),
    ('Dünnfilmwiderstände - SMD', 'Thin Film Resistors - SMD'),
    ('Widerstandnetzwerke u. -Arrays', 'Resistor Networks & Arrays'),
    ('Strommesswiderstände - SMD', 'Current Sense Resistors - SMD'),
    ('HF-Induktivitäten – SMD', 'RF Inductors - SMD'),
    ('HF-Induktivitäten - SMD', 'RF Inductors - SMD'),
    ('Lineare Spannungsregler', 'Linear Voltage Regulators'),
    ('Schaltspannungsregler', 'Switching Voltage Regulators'),
    ('HF-Transceiver', 'RF Transceivers'),
    ('HF-Schalt-ICs', 'RF Switch ICs'),
    ('Batterie-Management', 'Battery Management'),
    ('LED-Anzeigetreiber', 'LED Display Drivers'),
    ('Puffer u. Leitungstreiber', 'Buffers & Line Drivers'),
    ('Signalaufbereitung', 'Signal Conditioning'),
    ('ESD-Entstörer/TVS-Dioden', 'ESD Suppressors / TVS Diodes'),
    ('ESD-Entstörer', 'ESD Suppressors'),
    ('Dioden (Allzweck, Leistung, Schaltung)', 'Diodes (General Purpose, Power, Switching)'),
    ('Sensorschalter', 'Tactile Switches'),
    ('Infrarot-Receiver', 'Infrared Receivers'),
    ('Piezo-Summer und Audioanzeigen', 'Piezo Buzzers & Audio Indicators'),
    ('Speicherkartenverbinder', 'Memory Card Connectors'),
    ('USB-Stecker', 'USB Connectors'),
    ('FFC & FPC-Steckverbinder', 'FFC & FPC Connectors'),
    ('Sockel & Kabelgehäuse', 'Headers & Wire Housings'),
    ('Quarze', 'Crystals'),
    ('Sicherungen zur Aufbaumontage', 'Surface Mount Fuses'),
    ('Standard-LEDs - SMD', 'Standard LEDs - SMD'),
    ('Standard-LEDs', 'Standard LEDs'),
    ('Bipolartransistoren - BJT', 'Bipolar Transistors - BJT'),
    ('Antennen', 'Antennas'),
    # Lifecycle terms
    ('End-of-Life-Produkt', 'End of Life'),
    ('Beschränkte Verfügbarkeit', 'Limited Availability'),
    ('Neues Produkt', 'New Product'),
    # RoHS
    ('RoHS-konform', 'RoHS Compliant'),
    ('RoHS: Konform durch Ausnahmegenehmigung', 'RoHS Compliant (Exemption)'),
]


def translate_description(desc):
    """Translate a German Mouser description to English.

    Descriptions follow: "German Category  ENGLISH/UNIVERSAL SPECS"
    We translate the German prefix and keep the technical specs as-is.
    """
    result = desc
    for german, english in GERMAN_TRANSLATIONS:
        if german in result:
            result = result.replace(german, english)
    return result


def translate_lifecycle(lifecycle):
    """Translate German lifecycle status to English."""
    mapping = {
        'Neues Produkt': 'New Product',
        'End-of-Life-Produkt': 'End of Life',
        'Beschränkte Verfügbarkeit': 'Limited Availability',
        'NRND': 'Not Recommended for New Designs',
    }
    return mapping.get(lifecycle.strip(), lifecycle.strip())


def translate_rohs(rohs):
    """Translate German RoHS status to English."""
    mapping = {
        'RoHS-konform': 'RoHS Compliant',
        'RoHS: Konform durch Ausnahmegenehmigung': 'RoHS Compliant (Exemption)',
    }
    return mapping.get(rohs.strip(), rohs.strip())


class Command(BaseCommand):
    help = "Import Flipper Zero BOM data from Excel files into the database."

    def add_arguments(self, parser):
        parser.add_argument(
            "--flush", action="store_true",
            help="Delete existing BOM-sourced ProductComponent links before re-importing.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Show what would be imported without writing to the database.",
        )
        parser.add_argument(
            "--spreadsheet", action="store_true",
            help="Only generate the spreadsheet, don't import components.",
        )
        parser.add_argument(
            "--designators", action="store_true",
            help="Apply reference designators from schematic data to existing records.",
        )

    def handle(self, *args, **options):
        from apps.products.models import Product
        from apps.components.models import Component, ProductComponent

        try:
            product = Product.objects.get(manufacturer="Flipper Devices")
        except Product.DoesNotExist:
            self.stderr.write(self.style.ERROR(
                "Flipper Zero product not found. Run seed_data first."
            ))
            return

        # --designators only: apply reference designators and exit
        if options.get("designators"):
            self._apply_designators(product)
            return

        try:
            import xlrd
        except ImportError:
            self.stderr.write(self.style.ERROR(
                "xlrd is required: pip install xlrd"
            ))
            return

        admin = User.objects.filter(is_superuser=True).order_by("date_joined").first()
        if not admin:
            self.stderr.write(self.style.ERROR("No admin user found."))
            return

        base_dir = Path(settings.MEDIA_ROOT) / "downloads" / "flipper-zero"

        # Parse BOM files
        all_parts = []
        for filename, board_name in BOM_FILES:
            filepath = base_dir / filename
            if not filepath.exists():
                self.stderr.write(self.style.WARNING(f"File not found: {filepath}"))
                continue

            wb = xlrd.open_workbook(str(filepath))
            sheet = wb.sheet_by_index(0)

            for r in range(DATA_START, sheet.nrows):
                mfr_pn = str(sheet.cell_value(r, COL_MFR_PN)).strip()
                if not mfr_pn:
                    continue

                mfr = str(sheet.cell_value(r, COL_MFR)).strip()
                desc = str(sheet.cell_value(r, COL_DESC)).strip()
                mouser_pn = str(sheet.cell_value(r, COL_MOUSER)).strip()
                rohs = str(sheet.cell_value(r, COL_ROHS)).strip()
                lifecycle = str(sheet.cell_value(r, COL_LIFECYCLE)).strip()
                qty_raw = sheet.cell_value(r, COL_QTY)
                qty = int(qty_raw) if isinstance(qty_raw, float) else 1
                price = str(sheet.cell_value(r, COL_PRICE)).strip()
                total = str(sheet.cell_value(r, COL_TOTAL)).strip()

                ctype = classify_component(desc)
                package = extract_package(desc)
                value = extract_value(desc, ctype)
                function = translate_function(desc, ctype)
                specs = extract_specifications(desc, ctype)
                desc_en = translate_description(desc)
                rohs_en = translate_rohs(rohs)
                lifecycle_en = translate_lifecycle(lifecycle)

                all_parts.append({
                    'board': board_name,
                    'part_number': mfr_pn,
                    'manufacturer': mfr,
                    'description': desc_en,
                    'description_de': desc,
                    'mouser_pn': mouser_pn,
                    'rohs': rohs_en,
                    'lifecycle': lifecycle_en,
                    'quantity': qty,
                    'unit_price_eur': price,
                    'total_price_eur': total,
                    'component_type': ctype,
                    'package_type': package,
                    'primary_value': value,
                    'typical_function': function,
                    'specifications': specs,
                })

        self.stdout.write(f"Parsed {len(all_parts)} BOM entries from {len(BOM_FILES)} files.")

        # Always generate spreadsheet
        self._generate_spreadsheet(all_parts, product, base_dir)

        if options["spreadsheet"]:
            return

        if options["dry_run"]:
            self._dry_run(all_parts, product)
            return

        # Import into database
        self._import(all_parts, product, admin, options["flush"])

    def _dry_run(self, parts, product):
        from apps.components.models import Component

        new_components = 0
        existing_components = 0
        for p in parts:
            exists = Component.objects.filter(
                manufacturer=p['manufacturer'],
                part_number=p['part_number'],
            ).exists()
            if exists:
                existing_components += 1
            else:
                new_components += 1
                self.stdout.write(
                    f"  NEW: {p['manufacturer']} {p['part_number']} "
                    f"({p['component_type']}) - {p['typical_function']}"
                )

        self.stdout.write(f"\nDry run summary:")
        self.stdout.write(f"  Would create {new_components} new components")
        self.stdout.write(f"  Would reuse {existing_components} existing components")
        self.stdout.write(f"  Would create {len(parts)} ProductComponent links")

    @transaction.atomic
    def _import(self, parts, product, admin, flush):
        from apps.components.models import Component, ProductComponent

        if flush:
            # Only delete BOM-sourced links (basic submission level)
            count = ProductComponent.objects.filter(
                product=product,
                notes__startswith="[BOM Import]",
            ).count()
            ProductComponent.objects.filter(
                product=product,
                notes__startswith="[BOM Import]",
            ).delete()
            self.stdout.write(self.style.WARNING(
                f"Flushed {count} existing BOM-imported links."
            ))

        created_components = 0
        reused_components = 0
        created_links = 0
        skipped_links = 0

        for p in parts:
            # Get or create component
            component, was_created = Component.objects.get_or_create(
                manufacturer=p['manufacturer'],
                part_number=p['part_number'],
                defaults={
                    'component_type': p['component_type'],
                    'package_type': p['package_type'],
                    'typical_function': p['typical_function'],
                    'description': p['description'],
                    'specifications': p['specifications'],
                    'is_verified': True,
                    'created_by': admin,
                },
            )

            if was_created:
                created_components += 1
                self.stdout.write(
                    f"  Created: {p['manufacturer']} {p['part_number']} "
                    f"({p['component_type']})"
                )
            else:
                reused_components += 1
                # Update fields if empty or still in German
                updated = False
                if not component.package_type and p['package_type']:
                    component.package_type = p['package_type']
                    updated = True
                if not component.typical_function and p['typical_function']:
                    component.typical_function = p['typical_function']
                    updated = True
                # Replace German descriptions with English
                if component.description and any(
                    de in component.description for de, _ in GERMAN_TRANSLATIONS
                ):
                    component.description = p['description']
                    updated = True
                # Populate specifications if empty or merge new keys
                if p['specifications']:
                    if not component.specifications:
                        component.specifications = p['specifications']
                        updated = True
                    else:
                        # Merge: add keys from BOM that aren't already set
                        for k, v in p['specifications'].items():
                            if k not in component.specifications:
                                component.specifications[k] = v
                                updated = True
                if updated:
                    component.save()

            # Create ProductComponent link (skip if exact match exists)
            link_exists = ProductComponent.objects.filter(
                product=product,
                component=component,
                board_name=p['board'],
            ).exists()

            if link_exists:
                skipped_links += 1
                continue

            # Look up reference designator from schematic data
            ref_desig = REFERENCE_DESIGNATORS.get(p['part_number'], '')
            if isinstance(ref_desig, dict):
                ref_desig = ref_desig.get(p['board'], '')

            ProductComponent.objects.create(
                product=product,
                component=component,
                quantity=p['quantity'],
                board_name=p['board'],
                reference_designator=ref_desig,
                notes=f"[BOM Import] Mouser: {p['mouser_pn']}",
                submission_level='advanced',
                is_verified=True,
                created_by=admin,
            )
            created_links += 1

        self.stdout.write(f"\nImport complete:")
        self.stdout.write(f"  Components: {created_components} created, {reused_components} reused")
        self.stdout.write(f"  Links: {created_links} created, {skipped_links} already existed")

        total = ProductComponent.objects.filter(product=product).count()
        self.stdout.write(self.style.SUCCESS(
            f"Flipper Zero now has {total} component links."
        ))

    def _apply_designators(self, product):
        """Apply reference designator data from schematics to ProductComponent records."""
        from apps.components.models import ProductComponent

        pcs = ProductComponent.objects.filter(
            product=product
        ).select_related('component')

        updated = 0
        skipped = 0
        not_found = 0

        for pc in pcs:
            pn = pc.component.part_number
            mapping = REFERENCE_DESIGNATORS.get(pn)
            if mapping is None:
                not_found += 1
                continue

            # Resolve board-specific mappings
            if isinstance(mapping, dict):
                designators = mapping.get(pc.board_name, '')
            else:
                designators = mapping

            if not designators:
                not_found += 1
                continue

            if pc.reference_designator == designators:
                skipped += 1
                continue

            old = pc.reference_designator or '(empty)'
            pc.reference_designator = designators
            pc.save(update_fields=['reference_designator'])
            updated += 1
            self.stdout.write(
                f"  {pc.component.part_number} [{pc.board_name}]: "
                f"{old} → {designators}"
            )

        self.stdout.write(f"\nDesignator update complete:")
        self.stdout.write(f"  Updated: {updated}")
        self.stdout.write(f"  Already correct: {skipped}")
        self.stdout.write(f"  No mapping found: {not_found}")

        total_with = ProductComponent.objects.filter(
            product=product
        ).exclude(reference_designator='').count()
        total = ProductComponent.objects.filter(product=product).count()
        self.stdout.write(self.style.SUCCESS(
            f"  {total_with}/{total} links now have reference designators."
        ))

    def _generate_spreadsheet(self, parts, product, base_dir):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.stderr.write(self.style.WARNING(
                "openpyxl not installed — generating CSV instead."
            ))
            self._generate_csv(parts, product, base_dir)
            return

        wb = Workbook()

        # --- Sheet 1: Full BOM ---
        ws = wb.active
        ws.title = "Full BOM"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        board_fill_main = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        board_fill_ibtn = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Title rows
        ws.merge_cells('A1:P1')
        ws['A1'] = f"Flipper Zero ({product.model_number}) — Complete Bill of Materials"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:P2')
        ws['A2'] = "Source: Community DIY BOM (Mouser) | Generated by junkbin.io"
        ws['A2'].font = Font(italic=True, size=10, color="666666")

        # Headers
        headers = [
            ('Board', 15),
            ('Ref Designator', 20),
            ('Manufacturer', 22),
            ('Part Number', 30),
            ('Type', 14),
            ('Package', 10),
            ('Value', 14),
            ('Function', 35),
            ('Qty', 5),
            ('Mouser P/N', 22),
            ('Unit Price (EUR)', 14),
            ('Total (EUR)', 12),
            ('RoHS', 12),
            ('Lifecycle', 18),
            ('Description', 60),
            ('Description (German)', 60),
        ]

        for col, (name, width) in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        for i, p in enumerate(parts, 5):
            # Look up reference designator
            ref_desig = REFERENCE_DESIGNATORS.get(p['part_number'], '')
            if isinstance(ref_desig, dict):
                ref_desig = ref_desig.get(p['board'], '')

            row_data = [
                p['board'],
                ref_desig,
                p['manufacturer'],
                p['part_number'],
                p['component_type'],
                p['package_type'],
                p['primary_value'],
                p['typical_function'],
                p['quantity'],
                p['mouser_pn'],
                p['unit_price_eur'],
                p['total_price_eur'],
                p['rohs'],
                p['lifecycle'],
                p['description'],
                p['description_de'],
            ]
            fill = board_fill_main if p['board'] == 'Main PCB' else board_fill_ibtn
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.fill = fill
                if col == 9:  # Qty column
                    cell.alignment = Alignment(horizontal='center')

        # Freeze panes
        ws.freeze_panes = 'A5'

        # Auto-filter
        ws.auto_filter.ref = f"A4:P{4 + len(parts)}"

        # --- Sheet 2: Summary by Type ---
        ws2 = wb.create_sheet("Summary by Type")
        ws2.merge_cells('A1:E1')
        ws2['A1'] = "Component Summary by Type"
        ws2['A1'].font = Font(bold=True, size=14)

        sum_headers = ['Component Type', 'Line Items', 'Total Qty', 'Boards']
        for col, name in enumerate(sum_headers, 1):
            cell = ws2.cell(row=3, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        from collections import Counter, defaultdict
        type_counts = Counter()
        type_qty = Counter()
        type_boards = defaultdict(set)
        for p in parts:
            type_counts[p['component_type']] += 1
            type_qty[p['component_type']] += p['quantity']
            type_boards[p['component_type']].add(p['board'])

        for i, (ctype, count) in enumerate(type_counts.most_common(), 4):
            ws2.cell(row=i, column=1, value=ctype).border = thin_border
            ws2.cell(row=i, column=2, value=count).border = thin_border
            ws2.cell(row=i, column=3, value=type_qty[ctype]).border = thin_border
            ws2.cell(row=i, column=4, value=', '.join(sorted(type_boards[ctype]))).border = thin_border

        total_row = 4 + len(type_counts)
        ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws2.cell(row=total_row, column=2, value=len(parts)).font = Font(bold=True)
        ws2.cell(row=total_row, column=3, value=sum(p['quantity'] for p in parts)).font = Font(bold=True)

        for col in range(1, 5):
            ws2.column_dimensions[get_column_letter(col)].width = [20, 12, 12, 30][col - 1]

        # --- Sheet 3: Active Components Only ---
        ws3 = wb.create_sheet("Active Components")
        ws3.merge_cells('A1:G1')
        ws3['A1'] = "Active / Notable Components (ICs, Regulators, Connectors, Sensors)"
        ws3['A1'].font = Font(bold=True, size=14)

        active_headers = ['Board', 'Manufacturer', 'Part Number', 'Type', 'Function', 'Qty', 'Mouser P/N']
        active_types = {'ic', 'regulator', 'rf_module', 'sensor', 'connector', 'crystal', 'antenna', 'mosfet', 'transistor'}
        active_parts = [p for p in parts if p['component_type'] in active_types]

        for col, name in enumerate(active_headers, 1):
            cell = ws3.cell(row=3, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for i, p in enumerate(active_parts, 4):
            row = [p['board'], p['manufacturer'], p['part_number'],
                   p['component_type'], p['typical_function'],
                   p['quantity'], p['mouser_pn']]
            for col, val in enumerate(row, 1):
                cell = ws3.cell(row=i, column=col, value=val)
                cell.border = thin_border

        for col, w in enumerate([15, 22, 30, 12, 35, 6, 22], 1):
            ws3.column_dimensions[get_column_letter(col)].width = w

        # Save
        output_dir = Path(settings.MEDIA_ROOT) / "downloads" / "flipper-zero"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "flipper-zero-complete-bom.xlsx"
        wb.save(str(output_path))
        self.stdout.write(self.style.SUCCESS(
            f"Spreadsheet saved: {output_path}"
        ))

    def _generate_csv(self, parts, product, base_dir):
        """Fallback CSV generation if openpyxl is not available."""
        import csv

        output_dir = Path(settings.MEDIA_ROOT) / "downloads" / "flipper-zero"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "flipper-zero-complete-bom.csv"

        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Board', 'Ref Designator', 'Manufacturer', 'Part Number',
                'Type', 'Package', 'Value', 'Function', 'Qty', 'Mouser P/N',
                'Unit Price (EUR)', 'Total (EUR)', 'RoHS', 'Lifecycle',
                'Description', 'Description (German)',
            ])
            for p in parts:
                ref_desig = REFERENCE_DESIGNATORS.get(p['part_number'], '')
                if isinstance(ref_desig, dict):
                    ref_desig = ref_desig.get(p['board'], '')
                writer.writerow([
                    p['board'], ref_desig, p['manufacturer'],
                    p['part_number'], p['component_type'], p['package_type'],
                    p['primary_value'], p['typical_function'], p['quantity'],
                    p['mouser_pn'], p['unit_price_eur'], p['total_price_eur'],
                    p['rohs'], p['lifecycle'], p['description'],
                    p['description_de'],
                ])

        self.stdout.write(self.style.SUCCESS(f"CSV saved: {output_path}"))
